"""
Instagram投稿文生成アプリ
商品URLからクライアントごとのトンマナに合わせた投稿文を一括生成し、xlsxでダウンロード
"""

import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pdfplumber
import json
import os
import io
import re
import time
import base64
from datetime import datetime, timedelta, date as date_type
from pathlib import Path

# Google Generative AI
import google.generativeai as genai

# ── 設定 ──────────────────────────────────────────
CLIENTS_DIR = Path(__file__).parent / "clients"
CLIENTS_DIR.mkdir(exist_ok=True)

# GitHub API 永続化設定
GITHUB_TOKEN = st.secrets.get("GITHUB_TOKEN", "")
GITHUB_REPO = st.secrets.get("GITHUB_REPO", "fukudafukuo/instagram-caption-generator")
GITHUB_BRANCH = st.secrets.get("GITHUB_BRANCH", "main")
GITHUB_CLIENTS_DIR = "clients"
USE_GITHUB_STORAGE = bool(GITHUB_TOKEN)

st.set_page_config(
    page_title="Instagram投稿文ジェネレーター",
    page_icon="📸",
    layout="wide",
)

WEEKDAY_NAMES = ["月", "火", "水", "木", "金", "土", "日"]

# 投稿タイプ定義
POST_TYPES = {
    "single": "📷 単品紹介",
    "collection": "📸 集合カット（複数商品）",
    "brand": "💎 ブランドコンセプト",
}

# ── 季節イベント定義 ──────────────────────────────────
SEASONAL_EVENTS_BY_MONTH = {
    1: ["元旦・新年", "成人の日"],
    2: ["バレンタインデー", "節分"],
    3: ["ホワイトデー", "ひな祭り", "春分の日", "卒業・新生活準備"],
    4: ["イースター", "新生活シーズン", "花粉・ゆらぎ肌対策"],
    5: ["母の日", "ゴールデンウィーク", "紫外線対策"],
    6: ["父の日", "梅雨・湿気対策"],
    7: ["七夕", "夏本番・UV対策"],
    8: ["お盆", "夏バテ対策", "残暑ケア"],
    9: ["敬老の日", "秋分の日", "秋のスキンケア"],
    10: ["ハロウィン", "乾燥対策シーズン"],
    11: ["ブラックフライデー", "いい肌の日(11/8)"],
    12: ["クリスマス", "年末・冬の保湿ケア"],
}

ALL_EVENTS = []
for month, events in sorted(SEASONAL_EVENTS_BY_MONTH.items()):
    for ev in events:
        if ev not in ALL_EVENTS:
            ALL_EVENTS.append(ev)


def get_suggested_events(post_date):
    month = post_date.month
    events = list(SEASONAL_EVENTS_BY_MONTH.get(month, []))
    prev_month = 12 if month == 1 else month - 1
    next_month = 1 if month == 12 else month + 1
    for ev in SEASONAL_EVENTS_BY_MONTH.get(prev_month, []):
        if ev not in events:
            events.append(ev)
    for ev in SEASONAL_EVENTS_BY_MONTH.get(next_month, []):
        if ev not in events:
            events.append(ev)
    return events


# ── GitHub API ヘルパー ──────────────────────────────
def _gh_headers():
    return {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }


def _gh_get_file(filepath):
    """GitHub上のファイルを取得。(content_dict, error) を返す"""
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filepath}?ref={GITHUB_BRANCH}"
    resp = requests.get(url, headers=_gh_headers(), timeout=10)
    if resp.status_code == 200:
        return resp.json(), None
    elif resp.status_code == 404:
        return None, None
    else:
        return None, f"GitHub API error {resp.status_code}"


def _gh_put_file(filepath, content_bytes, message, sha=None):
    """GitHub上にファイルを作成/更新"""
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filepath}"
    body = {
        "message": message,
        "content": base64.b64encode(content_bytes).decode("ascii"),
        "branch": GITHUB_BRANCH,
    }
    if sha:
        body["sha"] = sha
    resp = requests.put(url, headers=_gh_headers(), json=body, timeout=10)
    return resp.status_code in (200, 201)


def _gh_delete_file(filepath, sha, message):
    """GitHub上のファイルを削除"""
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filepath}"
    body = {
        "message": message,
        "sha": sha,
        "branch": GITHUB_BRANCH,
    }
    resp = requests.delete(url, headers=_gh_headers(), json=body, timeout=10)
    return resp.status_code == 200


def _gh_list_dir(dirpath):
    """GitHub上のディレクトリ内ファイル一覧を取得"""
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{dirpath}?ref={GITHUB_BRANCH}"
    resp = requests.get(url, headers=_gh_headers(), timeout=10)
    if resp.status_code == 200:
        return resp.json(), None
    elif resp.status_code == 404:
        return [], None
    else:
        return [], f"GitHub API error {resp.status_code}"


# ── クライアントプロフィール管理 ──────────────────────
def load_client_list():
    if USE_GITHUB_STORAGE:
        files, err = _gh_list_dir(GITHUB_CLIENTS_DIR)
        if err:
            st.warning(f"クライアント一覧取得エラー: {err}")
            return {}
        clients = {}
        for f in files:
            if f["name"].endswith(".json"):
                cid = f["name"].replace(".json", "")
                # ファイル内容を取得して名前を読む
                file_data, _ = _gh_get_file(f"{GITHUB_CLIENTS_DIR}/{f['name']}")
                if file_data and "content" in file_data:
                    try:
                        raw = base64.b64decode(file_data["content"]).decode("utf-8")
                        data = json.loads(raw)
                        clients[cid] = data.get("name", cid)
                    except Exception:
                        clients[cid] = cid
                else:
                    clients[cid] = cid
        return clients
    else:
        clients = {}
        for f in CLIENTS_DIR.glob("*.json"):
            with open(f, "r", encoding="utf-8") as fp:
                data = json.load(fp)
                clients[f.stem] = data.get("name", f.stem)
        return clients


def load_client(client_id):
    if USE_GITHUB_STORAGE:
        filepath = f"{GITHUB_CLIENTS_DIR}/{client_id}.json"
        file_data, err = _gh_get_file(filepath)
        if err:
            st.warning(f"クライアント読込エラー: {err}")
            return None
        if file_data and "content" in file_data:
            try:
                raw = base64.b64decode(file_data["content"]).decode("utf-8")
                return json.loads(raw)
            except Exception:
                return None
        return None
    else:
        path = CLIENTS_DIR / f"{client_id}.json"
        if path.exists():
            with open(path, "r", encoding="utf-8") as fp:
                return json.load(fp)
        return None


def save_client(client_id, profile):
    if USE_GITHUB_STORAGE:
        filepath = f"{GITHUB_CLIENTS_DIR}/{client_id}.json"
        content = json.dumps(profile, ensure_ascii=False, indent=2).encode("utf-8")
        # 既存ファイルのSHAを取得（更新時に必要）
        existing, _ = _gh_get_file(filepath)
        sha = existing["sha"] if existing else None
        ok = _gh_put_file(filepath, content, f"Save client: {client_id}", sha)
        if not ok:
            st.error("クライアント保存に失敗しました。GitHub Token の権限を確認してください。")
    else:
        path = CLIENTS_DIR / f"{client_id}.json"
        with open(path, "w", encoding="utf-8") as fp:
            json.dump(profile, fp, ensure_ascii=False, indent=2)


def delete_client(client_id):
    if USE_GITHUB_STORAGE:
        filepath = f"{GITHUB_CLIENTS_DIR}/{client_id}.json"
        existing, _ = _gh_get_file(filepath)
        if existing:
            _gh_delete_file(filepath, existing["sha"], f"Delete client: {client_id}")
    else:
        path = CLIENTS_DIR / f"{client_id}.json"
        if path.exists():
            path.unlink()


def fetch_brand_concept(url, api_key):
    """ブランドサイトURLからページを取得し、Gemini APIでブランドコンセプトを要約する"""
    text, err = fetch_product_page(url)
    if err:
        return None, f"ページ取得エラー: {err}"
    if not text or len(text.strip()) < 50:
        return None, "ページから十分なテキストを取得できませんでした。"

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-2.5-flash")
        prompt = f"""以下はブランドの公式Webサイトのテキストです。
このブランドのコンセプト・理念・ストーリー・こだわりを300〜500文字程度で要約してください。
要約文のみを出力してください。前置きや説明は不要です。

【Webサイトのテキスト】
{text}
"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                response = model.generate_content(prompt)
                return response.text, None
            except Exception as e:
                if "429" in str(e) and attempt < max_retries - 1:
                    time.sleep(15 * (attempt + 1))
                else:
                    return None, f"AI要約エラー: {e}"
    except Exception as e:
        return None, f"API設定エラー: {e}"


def new_profile():
    return {
        "name": "",
        "brand_name": "",
        "brand_site_url": "",
        "brand_concept": "",
        "hashtag_fixed": "#美容好きな人と繋がりたい",
        "hashtag_limit": 5,
        "template": (
            "-———— -———— -————\n\n"
            "（ここにアカウント情報を入力）\n"
            "@アカウント名\n\n"
            "ブランドの紹介文をここに入力してください。\n\n"
            "製品のこだわりや詳細は、プロフィールのURLから\n"
            "公式HPをご覧ください☑️\n\n"
            "-———— -———— -————"
        ),
        "tone_instructions": (
            "・【見出し✨️】のような括弧付きヘッドラインで始める\n"
            "・「商品名は、」のように商品名を明示してから説明に入る\n"
            "・Instagram向けに1行15〜20文字程度で短く改行する\n"
            "・丁寧語を使用する\n"
            "・ポジティブな特徴の最後に◎を付ける\n"
            "・注釈は半角アスタリスク（*1, *2）を使用する"
        ),
        "sample_captions": "",
        "notes": (
            "・薬機法に抵触しないよう、商品ページに記載されている表現のみ使用すること\n"
            "・効果効能を断定する表現は避けること\n"
            "・商品ページのテキスト情報をベースに、表現を簡潔にまとめること"
        ),
    }


# ── 商品ページ取得 ──────────────────────────────────
def fetch_product_page(url):
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        }
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        resp.encoding = resp.apparent_encoding
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()
        main = soup.find("main") or soup.find("body")
        text = main.get_text(separator="\n", strip=True) if main else ""
        text = re.sub(r"\n{3,}", "\n\n", text)
        if len(text) > 8000:
            text = text[:8000] + "\n\n（以下省略）"
        return text, None
    except Exception as e:
        return None, str(e)


# ── リリース資料テキスト抽出（PDF / Excel）──────────
def extract_text_from_pdf(uploaded_file):
    """アップロードされたPDFファイルからテキストを抽出する"""
    try:
        pdf_bytes = uploaded_file.read()
        uploaded_file.seek(0)
        text_parts = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
                # テーブルがあれば抽出
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        cells = [str(c) if c else "" for c in row]
                        text_parts.append(" | ".join(cells))
        text = "\n\n".join(text_parts)
        text = re.sub(r"\n{3,}", "\n\n", text)
        if len(text) > 8000:
            text = text[:8000] + "\n\n（以下省略）"
        return text, None
    except Exception as e:
        return None, f"PDF読み取りエラー: {e}"


def extract_text_from_excel(uploaded_file):
    """アップロードされたExcelファイルからテキストを抽出する"""
    try:
        excel_bytes = uploaded_file.read()
        uploaded_file.seek(0)
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"【シート: {sheet_name}】")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells).strip()
                if line and line != " | " * (len(cells) - 1):
                    text_parts.append(line)
        text = "\n".join(text_parts)
        text = re.sub(r"\n{3,}", "\n\n", text)
        if len(text) > 8000:
            text = text[:8000] + "\n\n（以下省略）"
        return text, None
    except Exception as e:
        return None, f"Excel読み取りエラー: {e}"


def extract_text_from_file(uploaded_file):
    """ファイル種別に応じてテキスト抽出を振り分ける"""
    name = uploaded_file.name.lower()
    if name.endswith(".pdf"):
        return extract_text_from_pdf(uploaded_file)
    elif name.endswith((".xlsx", ".xls")):
        return extract_text_from_excel(uploaded_file)
    else:
        return None, f"未対応のファイル形式です: {name}"


# ── 投稿スケジュール生成（曜日ベース）────────────────
def generate_schedule_weekday(total_posts, start_date, post_weekdays):
    dates = []
    current = start_date
    while current.weekday() not in post_weekdays:
        current += timedelta(days=1)
    while len(dates) < total_posts:
        if current.weekday() in post_weekdays:
            dates.append(current)
        current += timedelta(days=1)
    return dates


# ── 投稿割り当て生成 ──────────────────────────────
def build_assignments(product_entries):
    """
    product_entries: list of dict
      - type: "single" | "collection" | "brand"
      - url: str (single用)
      - urls: str (collection用、改行区切り)
      - description: str (collection/brand用の補足)
      - count: int
    各エントリをcount回分、ラウンドロビンで投稿枠に割り当てる。
    返り値: list of dict (各投稿枠の情報)
    """
    remaining = []
    for entry in product_entries:
        remaining.append({"entry": entry, "left": entry["count"]})

    assignments = []
    total = sum(e["count"] for e in product_entries)
    while len(assignments) < total:
        for item in remaining:
            if item["left"] > 0:
                assignments.append(item["entry"])
                item["left"] -= 1
                if len(assignments) >= total:
                    break
    return assignments


# ── キャプション生成（Gemini API）──────────────────
def generate_caption(entry, product_texts, profile, api_key,
                     post_number=None, total_posts=None,
                     seasonal_event=None, post_date=None,
                     same_product_variation=None):
    """
    entry: 投稿エントリ情報 (type, url, urls, description, count)
    product_texts: dict of {url: text} 取得済みページテキスト
    """
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    post_type = entry.get("type", "single")

    # ── 共通プロンプト ──
    prompt = f"""あなたはInstagramの投稿文ライターです。
指定されたトンマナに合わせてInstagram投稿文を作成してください。
投稿文のみを出力してください。説明や前置きは不要です。

【ブランド名】
{profile.get('brand_name', '')}

【トンマナ指示】
{profile.get('tone_instructions', '')}

【注意事項】
{profile.get('notes', '')}

【ハッシュタグルール】
- 固定ハッシュタグ: {profile.get('hashtag_fixed', '')}
- ハッシュタグ上限: {profile.get('hashtag_limit', 5)}個
- ブランド名のハッシュタグを含めてください

【テンプレート（キャプション末尾に必ずこの定型文を付加してください）】
{profile.get('template', '')}

"""

    # ── タイプ別指示 ──
    input_method = entry.get("input_method", "url")

    if post_type == "single":
        prompt += """【投稿タイプ: 単品紹介】
1つの商品にフォーカスした投稿文を作成してください。
商品名のハッシュタグも含めてください。

"""
        if input_method == "file":
            file_text = entry.get("file_text", "")
            pname_manual = entry.get("product_name_manual", "")
            if pname_manual:
                prompt += f"【商品名】\n{pname_manual}\n\n"
            prompt += f"""【リリース資料からの商品情報】
{file_text}
"""
        else:
            url = entry.get("url", "")
            text = product_texts.get(url, "")
            prompt += f"""【商品ページ情報】
URL: {url}

{text}
"""

    elif post_type == "collection":
        desc = entry.get("description", "").strip()
        prompt += f"""【投稿タイプ: 集合カット（複数商品）】
写真には複数の商品が写っています。
ラインナップの魅力やスキンケアルーティンとしての使い方を紹介してください。
個々の商品を簡潔に紹介しつつ、組み合わせて使うメリットや全体の統一感を訴求してください。
"""
        if desc:
            prompt += f"""【写真の説明・切り口】
{desc}

"""
        if input_method == "file":
            file_text = entry.get("file_text", "")
            prompt += f"""【リリース資料からの商品情報】
{file_text}
"""
        else:
            urls_text = entry.get("urls", "")
            url_list = [u.strip() for u in urls_text.strip().split("\n") if u.strip()]
            for j, url in enumerate(url_list):
                text = product_texts.get(url, "")
                if text:
                    prompt += f"""【商品{j+1} ページ情報】
URL: {url}

{text}

"""

    elif post_type == "brand":
        desc = entry.get("description", "").strip()
        brand_concept = profile.get("brand_concept", "").strip()
        prompt += f"""【投稿タイプ: ブランドコンセプト】
ブランド全体のコンセプト、世界観、こだわりを紹介する投稿文を作成してください。
特定の商品名ではなく、ブランドとしての価値観・ストーリーを伝えてください。
"""
        if brand_concept:
            prompt += f"""【ブランドコンセプト情報】
{brand_concept}

"""
        if desc:
            prompt += f"""【投稿の切り口・テーマ】
{desc}

"""

    # ── 投稿番号 ──
    if post_number is not None and total_posts is not None:
        prompt += f"""【投稿位置】
この投稿は全{total_posts}投稿中の第{post_number}投稿目です。
"""

    # ── バリエーション ──
    if same_product_variation is not None and same_product_variation > 1:
        if post_type == "brand":
            prompt += f"""【バリエーション指示】
ブランドコンセプト投稿の{same_product_variation}回目です。
前回とは異なる切り口で作成してください。
例: 1回目→ブランドストーリー、2回目→開発のこだわり、3回目→ユーザーへのメッセージ、4回目→ブランドの未来像
"""
        elif post_type == "collection":
            prompt += f"""【バリエーション指示】
この組み合わせの{same_product_variation}回目の投稿です。
前回とは異なる切り口で作成してください。
例: 1回目→ラインナップ紹介、2回目→使う順番・ルーティン、3回目→各商品の相乗効果、4回目→朝晩の使い分け
"""
        else:
            prompt += f"""【バリエーション指示】
この商品は複数回投稿されます。今回は{same_product_variation}回目の投稿です。
前回とは異なる切り口・訴求ポイントで作成してください。
例: 1回目→商品の特徴紹介、2回目→使い方・テクスチャー、3回目→成分のこだわり、4回目→口コミ風・体験レビュー風
"""

    # ── 季節イベント ──
    if seasonal_event and post_date:
        date_str = post_date.strftime("%m/%d")
        prompt += f"""【季節イベント連動】
投稿予定日: {date_str}
関連する季節イベント: {seasonal_event}
投稿文の冒頭や導入部分で、このイベント・季節感を自然に絡めてください。
ただし、商品/ブランド紹介がメインであることを忘れずに。
"""

    # ── サンプル ──
    sample = profile.get("sample_captions", "").strip()
    if sample:
        prompt += f"""【サンプル投稿文（このスタイル・トーンに合わせてください）】
{sample}

"""

    # リトライ処理（429 レートリミット対策）
    max_retries = 5
    for attempt in range(max_retries):
        try:
            response = model.generate_content(prompt)
            return response.text
        except Exception as e:
            if "429" in str(e) and attempt < max_retries - 1:
                wait = 15 * (attempt + 1)
                st.warning(f"⏳ レートリミット到達。{wait}秒待機後にリトライします... ({attempt+2}/{max_retries})")
                time.sleep(wait)
            else:
                raise


# ── xlsx生成（スプレッドシート転記用フォーマット）─────────
def create_xlsx_schedule(results, schedule_dates, client_label):
    wb = Workbook()

    # === シート1: 配信原稿（横並び・スプレッドシート互換）===
    ws = wb.active
    ws.title = "配信原稿"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    date_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    caption_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    url_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(vertical="top", wrap_text=True)
    body_font = Font(name="Yu Gothic", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.column_dimensions["A"].width = 20
    labels = {
        1: "投稿番号", 2: "投稿日", 3: "投稿タイプ", 4: "商品名",
        5: "商品URL（ストーリー用）", 6: "Instagram配信原稿", 7: "季節イベント",
    }
    for row_num, label in labels.items():
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = Font(name="Yu Gothic", bold=True, size=10, color="FFFFFF")
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    for i, item in enumerate(results):
        col = i + 2
        ws.column_dimensions[get_column_letter(col)].width = 35

        cell = ws.cell(row=1, column=col, value=i + 1)
        cell.font = body_font; cell.alignment = center_align; cell.border = thin_border

        if i < len(schedule_dates):
            d = schedule_dates[i]
            date_str = f"{d.month}月{d.day}日{WEEKDAY_NAMES[d.weekday()]}曜日"
        else:
            date_str = ""
        cell = ws.cell(row=2, column=col, value=date_str)
        cell.font = body_font; cell.alignment = center_align
        cell.border = thin_border; cell.fill = date_fill

        cell = ws.cell(row=3, column=col, value=item.get("post_type_label", ""))
        cell.font = body_font; cell.alignment = center_align; cell.border = thin_border

        cell = ws.cell(row=4, column=col, value=item.get("product_name", ""))
        cell.font = body_font; cell.alignment = center_align; cell.border = thin_border

        cell = ws.cell(row=5, column=col, value=item.get("url", ""))
        cell.font = body_font; cell.border = thin_border; cell.fill = url_fill

        cell = ws.cell(row=6, column=col, value=item.get("caption", ""))
        cell.font = body_font; cell.alignment = wrap_align
        cell.border = thin_border; cell.fill = caption_fill

        cell = ws.cell(row=7, column=col, value=item.get("seasonal_event", ""))
        cell.font = body_font; cell.alignment = center_align; cell.border = thin_border

    ws.row_dimensions[6].height = 300

    # === シート2: 一覧表 ===
    ws2 = wb.create_sheet("一覧表")
    list_headers = ["No.", "投稿日", "タイプ", "商品名", "商品URL", "季節イベント", "キャプション"]
    for col, h in enumerate(list_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(name="Yu Gothic", bold=True, size=10, color="FFFFFF")
        cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border

    for i, item in enumerate(results):
        row = i + 2
        ws2.cell(row=row, column=1, value=i + 1).font = body_font
        ws2.cell(row=row, column=1).border = thin_border
        if i < len(schedule_dates):
            d = schedule_dates[i]
            date_str = f"{d.month}月{d.day}日{WEEKDAY_NAMES[d.weekday()]}曜日"
        else:
            date_str = ""
        ws2.cell(row=row, column=2, value=date_str).font = body_font
        ws2.cell(row=row, column=2).border = thin_border
        ws2.cell(row=row, column=3, value=item.get("post_type_label", "")).font = body_font
        ws2.cell(row=row, column=3).border = thin_border
        ws2.cell(row=row, column=4, value=item.get("product_name", "")).font = body_font
        ws2.cell(row=row, column=4).border = thin_border
        ws2.cell(row=row, column=5, value=item.get("url", "")).font = body_font
        ws2.cell(row=row, column=5).border = thin_border
        ws2.cell(row=row, column=6, value=item.get("seasonal_event", "")).font = body_font
        ws2.cell(row=row, column=6).border = thin_border
        cell = ws2.cell(row=row, column=7, value=item.get("caption", ""))
        cell.font = body_font; cell.alignment = wrap_align; cell.border = thin_border

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 25
    ws2.column_dimensions["E"].width = 40
    ws2.column_dimensions["F"].width = 20
    ws2.column_dimensions["G"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════
#  メインUI
# ══════════════════════════════════════════════════
def main():
    st.title("📸 Instagram投稿文ジェネレーター")
    st.caption("商品URLを入力 → 一括で投稿文を生成 → xlsxでダウンロード → スプレッドシートに転記")

    # ── APIキー ──
    api_key = ""
    try:
        api_key = st.secrets["GEMINI_API_KEY"]
    except Exception:
        pass
    if not api_key or api_key == "your-gemini-api-key-here":
        st.warning("⚠️ `.streamlit/secrets.toml` にGemini APIキーを設定してください。")
        api_key = st.text_input(
            "または、ここにGemini APIキーを入力してください（一時利用）",
            type="password",
            help="https://aistudio.google.com/apikey から無料で取得できます",
        )
        if not api_key:
            st.stop()

    # ── サイドバー: クライアント管理 ──
    with st.sidebar:
        st.header("👤 クライアント設定")
        clients = load_client_list()
        options = ["（新規作成）"] + [f"{cid} — {name}" for cid, name in clients.items()]
        selected = st.selectbox("クライアントを選択", options)

        if selected == "（新規作成）":
            client_id = st.text_input("クライアントID（半角英数）", value="",
                                      help="保存用のID。例: toutvert, brand_abc")
            profile = new_profile()
        else:
            client_id = selected.split(" — ")[0]
            profile = load_client(client_id) or new_profile()

        st.divider()
        st.subheader("📝 トンマナ設定")
        profile["name"] = st.text_input("表示名", value=profile.get("name", ""))
        profile["brand_name"] = st.text_input("ブランド名", value=profile.get("brand_name", ""))

        # ── ブランドコンセプト自動取得 ──
        profile["brand_site_url"] = st.text_input(
            "ブランドサイトURL（任意）",
            value=profile.get("brand_site_url", ""),
            placeholder="https://www.example.com/about",
            help="入力後「🔍 自動取得」ボタンで、サイトからブランドコンセプトを自動要約します")

        if st.button("🔍 ブランドコンセプトを自動取得", use_container_width=True,
                     disabled=not profile.get("brand_site_url", "").strip()):
            brand_url = profile["brand_site_url"].strip()
            with st.spinner("ブランドサイトを解析中..."):
                concept, fetch_err = fetch_brand_concept(brand_url, api_key)
                if fetch_err:
                    st.error(f"❌ {fetch_err}")
                elif concept:
                    profile["brand_concept"] = concept
                    st.success("✅ ブランドコンセプトを自動取得しました。下のテキストエリアで編集も可能です。")
                    st.rerun()

        profile["brand_concept"] = st.text_area(
            "ブランドコンセプト",
            value=profile.get("brand_concept", ""),
            height=100,
            help="ブランドの理念・ストーリー・こだわり等。ブランドコンセプト投稿で使用されます。上のボタンで自動取得、または手動入力できます")
        profile["tone_instructions"] = st.text_area(
            "トーン・マナー指示", value=profile.get("tone_instructions", ""),
            height=150, help="投稿文のスタイルを指定してください")
        profile["sample_captions"] = st.text_area(
            "サンプル投稿文（承認済みの例）", value=profile.get("sample_captions", ""),
            height=200, help="過去に承認された投稿文を1〜3件貼り付けてください")
        profile["template"] = st.text_area(
            "テンプレート（末尾定型文）", value=profile.get("template", ""),
            height=150, help="キャプション末尾に必ず付加される定型文")

        col1, col2 = st.columns(2)
        with col1:
            profile["hashtag_fixed"] = st.text_input(
                "固定ハッシュタグ", value=profile.get("hashtag_fixed", ""))
        with col2:
            profile["hashtag_limit"] = st.number_input(
                "ハッシュタグ上限", min_value=1, max_value=30,
                value=profile.get("hashtag_limit", 5))

        profile["notes"] = st.text_area("注意事項", value=profile.get("notes", ""), height=100)

        st.divider()
        col_save, col_del = st.columns(2)
        with col_save:
            if st.button("💾 保存", use_container_width=True):
                if client_id:
                    save_client(client_id, profile)
                    st.success(f"「{profile['name'] or client_id}」を保存しました")
                    st.rerun()
                else:
                    st.error("クライアントIDを入力してください")
        with col_del:
            if client_id and client_id in clients:
                if st.button("🗑️ 削除", use_container_width=True):
                    delete_client(client_id)
                    st.success("削除しました")
                    st.rerun()

    # ══════════════════════════════════════════════
    #  STEP 1: 撮影プラン設定
    # ══════════════════════════════════════════════
    st.header("① 撮影プラン設定")

    plan_col1, plan_col2 = st.columns(2)
    with plan_col1:
        total_posts = st.selectbox(
            "合計投稿数", options=[8, 12, 16, 24], index=1,
            help="撮影サービスのプランに対応")
    with plan_col2:
        start_date = st.date_input(
            "初回投稿日", value=datetime.now().date() + timedelta(days=7),
            help="指定曜日でない場合、直近の該当曜日から開始します")

    if total_posts == 24:
        schedule_label = "月・水・金（週3回）"
        post_weekdays = [0, 2, 4]
    else:
        schedule_label = "月・金（週2回）"
        post_weekdays = [0, 4]

    st.info(f"📅 投稿スケジュール: **{schedule_label}**（祝日考慮なし）")

    schedule_dates = generate_schedule_weekday(total_posts, start_date, post_weekdays)
    st.caption(
        f"配信期間: {schedule_dates[0].month}/{schedule_dates[0].day}"
        f"({WEEKDAY_NAMES[schedule_dates[0].weekday()]})"
        f" 〜 {schedule_dates[-1].month}/{schedule_dates[-1].day}"
        f"({WEEKDAY_NAMES[schedule_dates[-1].weekday()]})")

    # ══════════════════════════════════════════════
    #  STEP 2: 投稿を登録
    # ══════════════════════════════════════════════
    st.header("② 投稿を登録")
    st.caption("投稿タイプを選び、URLと投稿回数を設定してください。")

    if "products" not in st.session_state:
        st.session_state["products"] = [
            {"type": "single", "url": "", "urls": "", "description": "",
             "count": 1, "input_method": "url", "file_text": "", "file_name": ""}
        ]

    # セッションに file_text がない既存エントリを補完
    for p in st.session_state["products"]:
        p.setdefault("input_method", "url")
        p.setdefault("file_text", "")
        p.setdefault("file_name", "")

    products = st.session_state["products"]

    items_to_remove = []
    for i, prod in enumerate(products):
        with st.container(border=True):
            top_col1, top_col2, top_col3 = st.columns([2.5, 1, 0.5])
            with top_col1:
                type_options = list(POST_TYPES.keys())
                type_labels = list(POST_TYPES.values())
                current_idx = type_options.index(prod["type"]) if prod["type"] in type_options else 0
                chosen_label = st.selectbox(
                    f"投稿タイプ", options=type_labels, index=current_idx,
                    key=f"type_{i}", label_visibility="collapsed")
                chosen_type = type_options[type_labels.index(chosen_label)]
                products[i]["type"] = chosen_type
            with top_col2:
                products[i]["count"] = st.number_input(
                    "投稿数", min_value=1, max_value=total_posts,
                    value=prod["count"], key=f"pcount_{i}")
            with top_col3:
                if len(products) > 1:
                    if st.button("✕", key=f"del_{i}", help="削除"):
                        items_to_remove.append(i)

            # タイプ別の入力フィールド
            if chosen_type == "single":
                input_method = st.radio(
                    "情報ソース", options=["🔗 URL", "📎 リリース資料（PDF/Excel）"],
                    key=f"input_method_{i}",
                    index=0 if prod.get("input_method") == "url" else 1,
                    horizontal=True, label_visibility="collapsed")
                products[i]["input_method"] = "url" if "URL" in input_method else "file"

                if products[i]["input_method"] == "url":
                    products[i]["url"] = st.text_input(
                        "商品URL", value=prod.get("url", ""),
                        key=f"url_{i}",
                        placeholder="https://www.example.com/product/123")
                else:
                    uploaded = st.file_uploader(
                        "リリース資料をアップロード",
                        type=["pdf", "xlsx", "xls"],
                        key=f"file_{i}",
                        help="新発売商品のリリース資料（PDF・Excel）を添付してください")
                    if uploaded is not None:
                        if uploaded.name != prod.get("file_name", ""):
                            with st.spinner(f"📄 {uploaded.name} を読み取り中..."):
                                text, err = extract_text_from_file(uploaded)
                                if err:
                                    st.error(f"❌ {err}")
                                else:
                                    products[i]["file_text"] = text
                                    products[i]["file_name"] = uploaded.name
                                    st.success(f"✅ {uploaded.name} から情報を抽出しました")
                    if prod.get("file_text"):
                        with st.expander(f"📄 抽出済みテキスト: {prod.get('file_name', '')}"):
                            st.text(prod["file_text"][:500] + ("..." if len(prod["file_text"]) > 500 else ""))
                    # 商品名を手動入力（ファイルの場合URLがないため）
                    products[i]["product_name_manual"] = st.text_input(
                        "商品名（必須）", value=prod.get("product_name_manual", ""),
                        key=f"pname_{i}",
                        placeholder="例: トゥヴェール ミネラルサンスクリーン")

            elif chosen_type == "collection":
                input_method = st.radio(
                    "情報ソース", options=["🔗 URL", "📎 リリース資料（PDF/Excel）"],
                    key=f"input_method_{i}",
                    index=0 if prod.get("input_method") == "url" else 1,
                    horizontal=True, label_visibility="collapsed")
                products[i]["input_method"] = "url" if "URL" in input_method else "file"

                if products[i]["input_method"] == "url":
                    products[i]["urls"] = st.text_area(
                        "商品URL（1行1つ・複数可）", value=prod.get("urls", ""),
                        key=f"urls_{i}", height=80,
                        placeholder="https://www.example.com/product/123\nhttps://www.example.com/product/456")
                else:
                    uploaded_files = st.file_uploader(
                        "リリース資料をアップロード（複数可）",
                        type=["pdf", "xlsx", "xls"],
                        key=f"files_{i}",
                        accept_multiple_files=True,
                        help="複数商品の資料をまとめてアップロードできます")
                    if uploaded_files:
                        all_texts = []
                        all_names = []
                        for uf in uploaded_files:
                            with st.spinner(f"📄 {uf.name} を読み取り中..."):
                                text, err = extract_text_from_file(uf)
                                if err:
                                    st.error(f"❌ {uf.name}: {err}")
                                else:
                                    all_texts.append(f"【資料: {uf.name}】\n{text}")
                                    all_names.append(uf.name)
                        if all_texts:
                            combined = "\n\n".join(all_texts)
                            if len(combined) > 8000:
                                combined = combined[:8000] + "\n\n（以下省略）"
                            products[i]["file_text"] = combined
                            products[i]["file_name"] = ", ".join(all_names)
                            st.success(f"✅ {len(all_names)}件のファイルから情報を抽出しました")
                    if prod.get("file_text"):
                        with st.expander(f"📄 抽出済み: {prod.get('file_name', '')}"):
                            st.text(prod["file_text"][:500] + ("..." if len(prod["file_text"]) > 500 else ""))

                products[i]["description"] = st.text_input(
                    "写真の説明（任意）", value=prod.get("description", ""),
                    key=f"desc_{i}",
                    placeholder="例: スキンケア3点ラインナップ、朝のルーティンセット")

            elif chosen_type == "brand":
                products[i]["description"] = st.text_input(
                    "投稿の切り口（任意）", value=prod.get("description", ""),
                    key=f"bdesc_{i}",
                    placeholder="例: ブランド誕生ストーリー、開発者の想い、サステナビリティ")

    if items_to_remove:
        for idx in sorted(items_to_remove, reverse=True):
            products.pop(idx)
        st.rerun()

    if st.button("＋ 投稿を追加"):
        products.append({"type": "single", "url": "", "urls": "", "description": "",
                         "count": 1, "input_method": "url", "file_text": "", "file_name": ""})
        st.rerun()

    # バリデーション
    sum_assigned = sum(p["count"] for p in products)
    valid = True
    for p in products:
        if p["type"] == "single":
            if p.get("input_method") == "url" and not p.get("url", "").strip():
                valid = False
            elif p.get("input_method") == "file" and not p.get("file_text", "").strip():
                valid = False
        elif p["type"] == "collection":
            if p.get("input_method") == "url" and not p.get("urls", "").strip():
                valid = False
            elif p.get("input_method") == "file" and not p.get("file_text", "").strip():
                valid = False

    if not valid:
        st.warning("⚠️ URLまたはリリース資料が未入力の項目があります。")

    if sum_assigned != total_posts:
        st.warning(
            f"⚠️ 投稿数の合計が **{sum_assigned}** です。"
            f"合計投稿数 **{total_posts}** と一致させてください。")
    elif valid:
        st.success(f"✅ {len(products)}件 × 合計 {sum_assigned} 投稿 — OK")

    # 割り当て生成
    assignments = build_assignments(products)

    # ══════════════════════════════════════════════
    #  STEP 3: 季節イベント設定（投稿ごと）
    # ══════════════════════════════════════════════
    st.header("③ 季節イベント設定")
    st.caption("特定の投稿に季節イベントを絡めたい場合、チェックを入れてイベントを選択してください。")

    post_events = []
    num_display = min(len(schedule_dates), len(assignments))

    for i in range(num_display):
        d = schedule_dates[i]
        entry = assignments[i]
        date_str = f"{d.month}/{d.day}({WEEKDAY_NAMES[d.weekday()]})"

        # 表示用ラベル
        pt = entry.get("type", "single")
        if pt == "single":
            if entry.get("input_method") == "file":
                label = entry.get("product_name_manual", "") or entry.get("file_name", "") or "📎 資料"
            else:
                label = entry.get("url", "")
                label = label.rstrip("/").split("/")[-1] if label else "—"
        elif pt == "collection":
            if entry.get("input_method") == "file":
                label = entry.get("description", "") or entry.get("file_name", "") or "📎 集合カット"
            else:
                label = entry.get("description", "") or "集合カット"
        else:
            label = entry.get("description", "") or "ブランド"
        if len(label) > 25:
            label = label[:25] + "…"

        type_icon = {"single": "📷", "collection": "📸", "brand": "💎"}.get(pt, "")

        suggested = get_suggested_events(d)
        col_check, col_date, col_prod, col_event = st.columns([0.5, 1.5, 2, 2.5])

        with col_check:
            enabled = st.checkbox("", key=f"ev_check_{i}", value=False,
                                  label_visibility="collapsed")
        with col_date:
            st.text(f"#{i+1} {date_str}")
        with col_prod:
            st.text(f"{type_icon} {label}")
        with col_event:
            if enabled:
                event_options = ["（なし）"] + suggested
                for ev in ALL_EVENTS:
                    if ev not in event_options:
                        event_options.append(ev)
                chosen = st.selectbox(
                    "イベント", options=event_options,
                    key=f"ev_select_{i}", label_visibility="collapsed")
                if chosen == "（なし）":
                    post_events.append((False, ""))
                else:
                    post_events.append((True, chosen))
            else:
                post_events.append((False, ""))

    # ══════════════════════════════════════════════
    #  STEP 4: 一括生成
    # ══════════════════════════════════════════════
    st.divider()
    can_generate = (sum_assigned == total_posts) and valid

    if st.button("✨ 一括生成", type="primary", use_container_width=True,
                 disabled=not can_generate):
        results = []
        progress = st.progress(0, text="生成準備中...")

        # 全URLを収集してページ取得（キャッシュ）— ファイルベースはスキップ
        all_urls = set()
        for entry in assignments:
            if entry.get("input_method") == "file":
                continue  # ファイルベースはURL取得不要
            pt = entry.get("type", "single")
            if pt == "single":
                url = entry.get("url", "").strip()
                if url:
                    all_urls.add(url)
            elif pt == "collection":
                for u in entry.get("urls", "").strip().split("\n"):
                    u = u.strip()
                    if u:
                        all_urls.add(u)

        all_urls = list(all_urls)
        page_cache = {}
        for i, url in enumerate(all_urls):
            progress.progress(
                i / (len(all_urls) + total_posts),
                text=f"商品ページを取得中 ({i+1}/{len(all_urls)}): {url[:50]}...")
            text, err = fetch_product_page(url)
            if err:
                st.error(f"❌ {url}: {err}")
                page_cache[url] = ""
            else:
                page_cache[url] = text

        # キャプション生成
        # エントリIDでバリエーションカウント
        variation_counter = {}

        for i, entry in enumerate(assignments):
            # エントリの識別キー
            pt = entry.get("type", "single")
            im = entry.get("input_method", "url")
            if pt == "single":
                if im == "file":
                    entry_key = f"single:file:{entry.get('file_name', '')}"
                else:
                    entry_key = f"single:{entry.get('url', '')}"
            elif pt == "collection":
                if im == "file":
                    entry_key = f"collection:file:{entry.get('file_name', '')}"
                else:
                    entry_key = f"collection:{entry.get('urls', '')}"
            else:
                entry_key = f"brand:{entry.get('description', '')}"

            variation_counter[entry_key] = variation_counter.get(entry_key, 0) + 1
            variation_num = variation_counter[entry_key]

            post_date = schedule_dates[i] if i < len(schedule_dates) else None

            seasonal_event = None
            if i < len(post_events) and post_events[i][0]:
                seasonal_event = post_events[i][1]

            # 商品名の特定
            if pt == "single":
                if im == "file":
                    pname = entry.get("product_name_manual", "") or entry.get("file_name", "") or "新商品"
                    display_url = ""
                else:
                    url = entry.get("url", "").strip()
                    text = page_cache.get(url, "")
                    lines = [l.strip() for l in text.split("\n") if l.strip()]
                    pname = lines[0][:50] if lines else "不明"
                    display_url = url
            elif pt == "collection":
                pname = entry.get("description", "") or "集合カット"
                if im == "file":
                    display_url = ""
                else:
                    urls_str = entry.get("urls", "")
                    first_url = urls_str.strip().split("\n")[0].strip() if urls_str.strip() else ""
                    display_url = first_url
            else:
                pname = entry.get("description", "") or "ブランドコンセプト"
                display_url = ""

            progress.progress(
                (len(all_urls) + i) / (len(all_urls) + total_posts),
                text=f"キャプション生成中 ({i+1}/{total_posts}): {pname}")

            try:
                caption = generate_caption(
                    entry, page_cache, profile, api_key,
                    post_number=i + 1, total_posts=total_posts,
                    seasonal_event=seasonal_event, post_date=post_date,
                    same_product_variation=variation_num)
            except Exception as e:
                st.error(f"❌ AI生成エラー ({pname}): {e}")
                caption = f"生成エラー: {e}"

            results.append({
                "url": display_url,
                "product_name": pname,
                "caption": caption,
                "seasonal_event": seasonal_event or "",
                "post_type_label": POST_TYPES.get(pt, ""),
            })

            # レートリミット回避
            if i < len(assignments) - 1:
                time.sleep(5)

        progress.progress(1.0, text="✅ 全投稿の生成が完了しました！")
        st.session_state["results"] = results
        st.session_state["schedule_dates"] = schedule_dates

    # ══════════════════════════════════════════════
    #  結果表示 & ダウンロード
    # ══════════════════════════════════════════════
    if "results" in st.session_state and st.session_state["results"]:
        results = st.session_state["results"]
        sched = st.session_state.get("schedule_dates", [])

        st.header("📝 生成結果（編集可能）")

        for i, item in enumerate(results):
            if i < len(sched):
                d = sched[i]
                date_label = f"{d.month}/{d.day}({WEEKDAY_NAMES[d.weekday()]})"
            else:
                date_label = ""

            event_label = f" 🎉{item.get('seasonal_event', '')}" if item.get("seasonal_event") else ""
            type_label = f" {item.get('post_type_label', '')}" if item.get("post_type_label") else ""

            with st.expander(
                f"**#{i+1} {date_label}**{type_label} — {item['product_name']}{event_label}",
                expanded=(i < 3)):
                edited = st.text_area(
                    "キャプション", value=item["caption"], height=400,
                    key=f"caption_{i}", label_visibility="collapsed")
                results[i]["caption"] = edited

        st.divider()
        st.subheader("📥 ダウンロード")
        st.caption(
            "**配信原稿シート**: 横並びフォーマット（スプレッドシートに新規タブとしてインポート → "
            "キャプション行をコピーして既存シートに貼り付け）\n\n"
            "**一覧表シート**: 縦並びフォーマット（確認・編集用）")

        xlsx_buf = create_xlsx_schedule(
            results, sched, profile.get("name") or client_id or "output")
        client_label = profile.get("name") or client_id or "output"
        filename = f"instagram_captions_{client_label}_{total_posts}posts.xlsx"

        st.download_button(
            label=f"📥 xlsxをダウンロード（{total_posts}投稿分）",
            data=xlsx_buf, file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True)


if __name__ == "__main__":
    main()
